import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import date, datetime

from tkcalendar import DateEntry

from db import (
    get_active_employees,
    add_attendance,
    get_conn,
    update_attendance as db_update_attendance,
    delete_attendance as db_delete_attendance,
)
from utils import month_date_range


class AttendanceTab(ttk.Frame):
    def __init__(self, master):
        super().__init__(master)

        self.employees_cache = []
        self.selected_att_id = None
        self.show_only_selected_var = tk.IntVar(value=0)

        self.build_ui()
        self.load_employees()
        self.load_current_month_logs()

    # ------------- ARAYÜZ ------------- #

    def build_ui(self):
        # Üst: Kayıt Formu
        form_frame = ttk.LabelFrame(self, text="Devamsızlık / Saat Düşme Kaydı")
        form_frame.pack(side="top", fill="x", padx=10, pady=10)

        # Personel
        ttk.Label(form_frame, text="Personel:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.att_emp_var = tk.StringVar()
        self.att_emp_combo = ttk.Combobox(form_frame, textvariable=self.att_emp_var, state="readonly", width=30)
        self.att_emp_combo.grid(row=0, column=1, padx=5, pady=5)
        self.att_emp_combo.bind("<<ComboboxSelected>>", self.on_employee_changed)

        # Personel Yenile butonu
        ttk.Button(form_frame, text="Yenile", command=self.refresh_employees).grid(
            row=0, column=2, padx=5, pady=5, sticky="w"
        )

        # Tarih (takvimli)
        ttk.Label(form_frame, text="Tarih:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.att_date_var = tk.StringVar()
        self.att_date_entry = DateEntry(
            form_frame,
            textvariable=self.att_date_var,
            date_pattern="yyyy-mm-dd",
            width=12
        )
        self.att_date_entry.set_date(date.today())
        self.att_date_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        # Tür
        ttk.Label(form_frame, text="Tür:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.att_type_var = tk.StringVar()
        self.att_type_combo = ttk.Combobox(
            form_frame,
            textvariable=self.att_type_var,
            state="readonly",
            values=[
                "HOUR_LOSS - Saatlik Eksik",
                "FULL_ABSENCE - Tam Gün Yok",
                "FREE_LEAVE - Ücretsiz İzin",
                "ANNUAL_LEAVE - Yıllık İzin",
                "REPORT - Rapor"
            ],
            width=30
        )
        self.att_type_combo.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        self.att_type_combo.current(0)

        # Eksik saat
        ttk.Label(form_frame, text="Eksik Saat:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.att_hours_var = tk.StringVar(value="0")
        self.att_hours_entry = ttk.Entry(form_frame, textvariable=self.att_hours_var, width=10)
        self.att_hours_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")

        # Not
        ttk.Label(form_frame, text="Not:").grid(row=4, column=0, padx=5, pady=5, sticky="w")
        self.att_note_var = tk.StringVar()
        ttk.Entry(form_frame, textvariable=self.att_note_var, width=40).grid(
            row=4, column=1, padx=5, pady=5, sticky="w"
        )

        # Butonlar
        btn_frame = ttk.Frame(form_frame)
        btn_frame.grid(row=5, column=0, columnspan=4, pady=10, sticky="w")

        ttk.Button(btn_frame, text="Yeni", command=self.clear_form).grid(row=0, column=0, padx=5)
        ttk.Button(btn_frame, text="Kaydet", command=self.save_attendance).grid(row=0, column=1, padx=5)
        ttk.Button(btn_frame, text="Sil", command=self.delete_selected).grid(row=0, column=2, padx=5)
        ttk.Button(btn_frame, text="Excel'den İçeri Aktar", command=self.import_from_excel).grid(row=0, column=3, padx=10)

        # Alt: Kayıt listesi
        list_frame = ttk.LabelFrame(self, text="Devamsızlık Kayıtları")
        list_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Filtre: sadece seçili personel
        top_filter = ttk.Frame(list_frame)
        top_filter.pack(side="top", anchor="w", padx=5, pady=(5, 0))
        ttk.Checkbutton(
            top_filter,
            text="Sadece seçili personel",
            variable=self.show_only_selected_var,
            command=self.load_current_month_logs
        ).pack(side="left")

        # Yıl / Ay filtresi
        period_frame = ttk.Frame(list_frame)
        period_frame.pack(side="top", anchor="w", padx=5, pady=(2, 5))
        ttk.Label(period_frame, text="Yıl:").pack(side="left")
        self.rep_year_var = tk.StringVar(value=str(date.today().year))
        ttk.Entry(period_frame, textvariable=self.rep_year_var, width=6).pack(side="left", padx=(0, 10))
        ttk.Label(period_frame, text="Ay:").pack(side="left")
        self.rep_month_var = tk.StringVar(value=str(date.today().month))
        ttk.Entry(period_frame, textvariable=self.rep_month_var, width=4).pack(side="left")
        ttk.Button(period_frame, text="Listele", command=self.load_current_month_logs).pack(side="left", padx=10)

        columns = ("date", "employee", "type", "hours", "note")
        self.att_tree = ttk.Treeview(list_frame, columns=columns, show="headings")
        self.att_tree.heading("date", text="Tarih")
        self.att_tree.heading("employee", text="Personel")
        self.att_tree.heading("type", text="Tür")
        self.att_tree.heading("hours", text="Saat")
        self.att_tree.heading("note", text="Not")

        self.att_tree.column("date", width=100)
        self.att_tree.column("employee", width=150)
        self.att_tree.column("type", width=130)
        self.att_tree.column("hours", width=60)
        self.att_tree.column("note", width=250)

        self.att_tree.pack(fill="both", expand=True)
        self.att_tree.bind("<<TreeviewSelect>>", self.on_select)

    # ------------- VERİ YÜKLEME ------------- #

    def clear_form(self):
        self.selected_att_id = None
        if self.att_emp_combo["values"]:
            self.att_emp_combo.current(0)
        self.att_date_entry.set_date(date.today())
        self.att_type_combo.current(0)
        self.att_hours_var.set("0")
        self.att_note_var.set("")

    def load_employees(self):
        employees = get_active_employees()
        self.employees_cache = employees
        names = [e[1] for e in employees]  # e = (id, name, rate, start_date)
        self.att_emp_combo["values"] = names
        if names:
            self.att_emp_combo.current(0)
        else:
            self.att_emp_var.set("")

    def refresh_employees(self):
        self.load_employees()
        messagebox.showinfo("Bilgi", "Personel listesi güncellendi.")
        self.load_current_month_logs()

    def on_employee_changed(self, event):
        if self.show_only_selected_var.get() == 1:
            self.load_current_month_logs()

    def load_current_month_logs(self):
        # Ağacı temizle
        for row in self.att_tree.get_children():
            self.att_tree.delete(row)

        # Seçili yıl/ay
        try:
            year = int(self.rep_year_var.get())
            month = int(self.rep_month_var.get())
            if not (1 <= month <= 12):
                raise ValueError
        except Exception:
            today = date.today()
            year, month = today.year, today.month
            self.rep_year_var.set(str(year))
            self.rep_month_var.set(str(month))

        month_start, month_end = month_date_range(year, month)

        # Filtre: sadece seçili personel mi?
        emp_filter_id = None
        if self.show_only_selected_var.get() == 1 and self.att_emp_var.get():
            selected_name = self.att_emp_var.get()
            for e in self.employees_cache:
                if e[1] == selected_name:
                    emp_filter_id = e[0]
                    break

        conn = get_conn()
        c = conn.cursor()
        query = """
            SELECT a.id, a.date, e.name, a.type, a.hours, a.note
            FROM attendance_logs a
            JOIN employees e ON e.id = a.employee_id
            WHERE a.date >= ? AND a.date < ?
        """
        params = [month_start.isoformat(), month_end.isoformat()]

        if emp_filter_id is not None:
            query += " AND a.employee_id = ?"
            params.append(emp_filter_id)

        query += " ORDER BY a.date DESC, e.name"

        c.execute(query, params)
        rows = c.fetchall()
        conn.close()

        for att_id, d_str, name, typ, hrs, note in rows:
            self.att_tree.insert(
                "",
                "end",
                iid=str(att_id),
                values=(d_str, name, typ, hrs, note or "")
            )

    # ------------- KAYIT SEÇİMİ ------------- #

    def on_select(self, event):
        item_id = self.att_tree.focus()
        if not item_id:
            return

        self.selected_att_id = int(item_id)
        vals = self.att_tree.item(item_id, "values")
        d_str, emp_name, typ, hours, note = vals

        self.att_emp_var.set(emp_name)

        self.att_date_var.set(d_str)
        try:
            self.att_date_entry.set_date(datetime.strptime(d_str, "%Y-%m-%d").date())
        except Exception:
            pass

        for i, v in enumerate(self.att_type_combo["values"]):
            if v.startswith(typ):
                self.att_type_combo.current(i)
                break

        if typ == "HOUR_LOSS":
            self.att_hours_var.set(str(hours))
        else:
            self.att_hours_var.set("0")

        self.att_note_var.set(note)

    # ------------- KAYDET / SİL ------------- #

    def save_attendance(self):
        emp_name = self.att_emp_var.get()
        if not emp_name:
            messagebox.showerror("Hata", "Lütfen personel seçin.")
            return

        emp_id = None
        for e in self.employees_cache:
            if e[1] == emp_name:
                emp_id = e[0]
                break

        if emp_id is None:
            messagebox.showerror("Hata", "Personel bulunamadı.")
            return

        date_str = self.att_date_var.get().strip()
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Hata", "Tarih formatı geçersiz.")
            return

        type_full = self.att_type_var.get()
        type_code = type_full.split(" - ")[0]

        hours_str = self.att_hours_var.get().strip() or "0"
        try:
            hours_val = float(hours_str.replace(",", "."))
        except ValueError:
            messagebox.showerror("Hata", "Eksik saat sayısal olmalıdır.")
            return

        if type_code != "HOUR_LOSS":
            hours_val = 0.0

        note = self.att_note_var.get().strip()

        if self.selected_att_id is None:
            add_attendance(emp_id, date_str, type_code, hours_val, note)
        else:
            db_update_attendance(self.selected_att_id, emp_id, date_str, type_code, hours_val, note)

        messagebox.showinfo("Başarılı", "Kayıt kaydedildi.")
        self.clear_form()
        self.load_current_month_logs()

    def delete_selected(self):
        if self.selected_att_id is None:
            messagebox.showwarning("Uyarı", "Silmek için listeden bir kayıt seçin.")
            return

        if not messagebox.askyesno("Onay", "Bu devamsızlık kaydını silmek istediğinize emin misiniz?"):
            return

        db_delete_attendance(self.selected_att_id)
        self.selected_att_id = None
        self.clear_form()
        self.load_current_month_logs()

    # ------------- EXCEL İÇE AKTAR (Önceki sürümün aynısı) ------------- #
    # Burayı değiştirmedim; sadece üstte rapor filtresini ekledik.

    def import_from_excel(self):
        """
        Satır bazlı devamsızlık Excel'ini içeri alır.
        Beklenen başlıklar:
        Personel | Tarih | Tür | Eksik Saat | Not
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            messagebox.showerror(
                "Hata",
                "Excel içe aktarma için 'openpyxl' kütüphanesi gerekiyor.\n\n"
                "Kurulum örneği:\n\npip install openpyxl"
            )
            return

        file_path = filedialog.askopenfilename(
            title="Devamsızlık Excel Dosyası Seç",
            filetypes=[("Excel Dosyası", "*.xlsx"), ("Tüm Dosyalar", "*.*")]
        )
        if not file_path:
            return

        self.load_employees()
        wb = None
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            messagebox.showerror("Hata", f"Excel dosyası açılamadı:\n{e}")
            return

        ws = wb.active

        header_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if not val:
                continue
            key = str(val).strip().lower()
            header_map[key] = col

        def find_col(possible_keys):
            for k in possible_keys:
                for h, col_idx in header_map.items():
                    if k in h:
                        return col_idx
            return None

        col_person = find_col(["personel", "isim", "ad"])
        col_date = find_col(["tarih", "date"])
        col_type = find_col(["tür", "tur", "type"])
        col_hours = find_col(["eksik saat", "saat", "hours"])
        col_note = find_col(["not", "açıklama", "aciklama", "note"])

        if not col_person or not col_date:
            messagebox.showerror(
                "Hata",
                "Excel başlıkları bulunamadı.\n\n"
                "En az şu başlıklar olmalı:\n"
                "- 'Personel'\n"
                "- 'Tarih'\n\n"
                "Opsiyonel: 'Tür', 'Eksik Saat', 'Not / Açıklama'"
            )
            return

        imported = 0
        skipped = 0
        unknown_emps = set()

        emp_name_to_id = {e[1]: e[0] for e in self.employees_cache}

        for row in range(2, ws.max_row + 1):
            name_val = ws.cell(row=row, column=col_person).value
            date_val = ws.cell(row=row, column=col_date).value

            if (name_val is None or str(name_val).strip() == "") and not date_val:
                continue

            name_str = str(name_val).strip() if name_val else ""
            if not name_str:
                skipped += 1
                continue

            emp_id = emp_name_to_id.get(name_str)
            if not emp_id:
                unknown_emps.add(name_str)
                skipped += 1
                continue

            date_str = None
            if isinstance(date_val, (datetime, date)):
                date_str = date_val.strftime("%Y-%m-%d")
            elif isinstance(date_val, str):
                dt = None
                for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
                    try:
                        dt = datetime.strptime(date_val.strip(), fmt)
                        break
                    except ValueError:
                        continue
                if dt is None:
                    skipped += 1
                    continue
                date_str = dt.strftime("%Y-%m-%d")
            else:
                skipped += 1
                continue

            type_code = "HOUR_LOSS"
            if col_type:
                type_val = ws.cell(row=row, column=col_type).value
                if type_val:
                    t_str = str(type_val).strip().upper()
                    if "HOUR" in t_str or "SAAT" in t_str:
                        type_code = "HOUR_LOSS"
                    elif "FULL" in t_str or "TAM" in t_str:
                        type_code = "FULL_ABSENCE"
                    elif "FREE" in t_str or "ÜCRETSİZ" in t_str:
                        type_code = "FREE_LEAVE"
                    elif "ANNUAL" in t_str or "YILLIK" in t_str:
                        type_code = "ANNUAL_LEAVE"
                    elif "RAPOR" in t_str or "REPORT" in t_str:
                        type_code = "REPORT"
                    elif t_str in ["HOUR_LOSS", "FULL_ABSENCE", "FREE_LEAVE", "ANNUAL_LEAVE", "REPORT"]:
                        type_code = t_str

            hours_val = 0.0
            if col_hours:
                h_val = ws.cell(row=row, column=col_hours).value
                if h_val is not None and str(h_val).strip() != "":
                    try:
                        hours_val = float(str(h_val).replace(",", "."))
                    except ValueError:
                        hours_val = 0.0

            if type_code != "HOUR_LOSS":
                hours_val = 0.0

            note_val = ""
            if col_note:
                n_val = ws.cell(row=row, column=col_note).value
                if n_val:
                    note_val = str(n_val).strip()

            try:
                add_attendance(emp_id, date_str, type_code, hours_val, note_val)
                imported += 1
            except Exception:
                skipped += 1

        self.load_current_month_logs()

        msg = f"İçe aktarma tamamlandı.\n\nBaşarılı kayıt: {imported}\nAtlanan satır: {skipped}"
        if unknown_emps:
            msg += "\n\nTanımlı olmayan personeller (eklemeniz gerekiyor):\n"
            msg += "\n".join(sorted(unknown_emps))
        messagebox.showinfo("Excel İçe Aktar", msg)
